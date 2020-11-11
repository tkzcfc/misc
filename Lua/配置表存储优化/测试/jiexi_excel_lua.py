# coding=utf-8 #
from openpyxl import load_workbook
import codecs
import sys
import os

column_separator = ', '
row_separator = '\n'
output_encoding = 'utf-8'
input_ext = '.xlsx'
input_ext2 = '.xlsm'
output_folder = 'output'

LANG_FILE = "output/LangXXXXX.lua"
NO_OPTIMIZE_FILE_LIST = [
    
]

def duqu_gongzuobo(path, name, ext, output_file):
    file_name = '%s/%s%s' % (path, name, ext)
    #print 'process xls: %s' % (file_name)
    workbook = load_workbook(file_name, True, False, True, True)
    sheet_names = workbook.get_sheet_names()
    for sheet_name in sheet_names:
        if not sheet_name.startswith('_'):
            output_file.write('_Config.%s = require("app.WDConfig.%s")' % (sheet_name, sheet_name))
            output_file.write(row_separator)
            sheet = workbook.get_sheet_by_name(sheet_name)
            output_name = dedao_shuchu_mingzi(sheet_name, sheet_name)
            #print 'output txt: %s' % (output_name)
            duqu_biao(sheet, output_name, name)
            # 
            if output_name in NO_OPTIMIZE_FILE_LIST:
                pass
            else:
                os.system("luajit.exe DataTableOptimizer.lua %s %s" % (output_name, output_name))
			

def dedao_shuchu_mingzi(workbook_name, worksheet_name):
    return '%s/%s.lua' % (output_folder, workbook_name)


def duqu_biao(sheet, filename, name):
    output_file = codecs.open(filename, 'w', output_encoding)
    cell_value = ""
    count_row = 0
    count_col = 0
    count_line = 0
    arrSkip = []
    rowFirstVlaue = []
    title = []
    count_row2 = 0
    count_skip = 10
    real_count = 0
    for row in sheet.iter_rows():
        count_col = 0
        for cell in row:
            value = cell.internal_value
            if count_col == 0:
                if value == None:
                    rowFirstVlaue.append(0)
                else:
                    rowFirstVlaue.append(1)
				
            if not isinstance(value, unicode):
                try:
                    cell_value = str(int(value))
                except:
                    cell_value = str(-1)
            else:
                cell_value = value
            if count_row == count_skip:
                title.append(cell_value)
                if cell_value.startswith('_') or cell_value.startswith('#'):
                    arrSkip.append(0)
                else:
                    if cell_value.startswith('i'):
                        arrSkip.append(1)
                    else:
                        if cell_value.startswith('f'):
                            arrSkip.append(2)
                        else:
                            arrSkip.append(3)
            count_col = count_col + 1
        count_row = count_row + 1
        #print 'rowFirstVlaue : %s' % (rowFirstVlaue)
		
    output_file.write("-- Auto Generate by Excel %s.xlsx, Don't try to modify!" % (name))
    output_file.write(row_separator)
    output_file.write(row_separator)
    output_file.write('local _Config = {}')
    output_file.write(row_separator)
    output_file.write(row_separator)
    output_file.write('_Config._Data = {}')
    output_file.write(row_separator)

    alldata = {}
    for row in sheet.iter_rows():
        rowkey = ""
        rowstr = ""
        if count_row2 < count_skip + 1:
            count_row2 = count_row2 + 1
            continue
			
        if rowFirstVlaue[count_row2] == 0:
            break
        
        if count_row2 > count_skip + 1:
            if filename != LANG_FILE:
                rowstr = rowstr + row_separator

        isFirstValue = 1
        real_count = real_count + 1
		
        count_line = 0
        for cell in row:
            if arrSkip[count_line] != 0:
                if count_line != 0:
                    rowstr = rowstr + column_separator
                value = cell.internal_value
                if not isinstance(value, unicode):
                    try:
                        if arrSkip[count_line] == 1:
                            cell_value = str(int(value))
                        elif arrSkip[count_line] == 2:
                            cell_value = str(float(value))
                        else:
                            cell_value = str(int(value))
                    except:
                        if arrSkip[count_line] == 1:
                            cell_value = str(-1)
                        else:
                            cell_value = str(-1)
                else:
                    if value == "":
                        if arrSkip[count_line] == 1:
                            cell_value = str(-1)                
                    else:
                        cell_value = value
                if arrSkip[count_line] == 3:
                    if cell_value == "nil":
                        if isFirstValue == 1:
                            rowstr = rowstr + '_Config._Data.id_%s = {' % (cell_value)
                            rowkey = cell_value
                            isFirstValue = 0
                        rowstr = rowstr + '%s = %s' % (title[count_line][1:], cell_value)
                    else:
                        if isFirstValue == 1:
                            rowstr = rowstr + '_Config._Data.id_%s = {' % (cell_value)
                            rowkey = cell_value
                            isFirstValue = 0
                        rowstr = rowstr + '%s = \'%s\'' % (title[count_line][1:], cell_value)
                else:
                    if isFirstValue == 1:
                        rowstr = rowstr + '_Config._Data.id_%s = {' % (cell_value)
                        rowkey = cell_value
                        isFirstValue = 0
                    rowstr = rowstr + '%s = %s' % (title[count_line][1:], cell_value)
            count_line = count_line + 1
        rowstr = rowstr + '}'
        if filename != LANG_FILE:
            output_file.write(rowstr)
        alldata[rowkey.zfill(8)] = rowstr
        cell_value = ""
        count_row2 = count_row2 + 1
    print filename
    if filename == LANG_FILE:
        currentid = 0
        output_lang_file = ""
        sorted_key_list = sorted(alldata)
        print len(sorted_key_list)
        for a in sorted_key_list:
            id = int(a)
            nowid = id/1500
            #print "    " + str(id) + " - " + str(nowid) + " - " + str(currentid)

            if output_lang_file != "" and nowid > currentid:
                #print "!!!!!!!!!!! " + str(id)
                output_lang_file.write("end" + row_separator)
                output_lang_file.write("return _LandConfig" + row_separator)
                output_lang_file.close()
                output_lang_file = ""

            if output_lang_file == "":
                output_file.write('require("app.WDConfig.Lang%d").AddData(_Config)' % (nowid) + row_separator)
				
                #print "new file" + str(nowid)
                currentid = nowid
                output_lang_file = codecs.open(filename.replace("Lang","Lang"+str(nowid)), 'w', output_encoding)
                output_lang_file.write("-- Auto Generate by Excel language.xlsx, Don't try to modify!" + row_separator)
                output_lang_file.write("local _LandConfig = {}" + row_separator)
                output_lang_file.write("function _LandConfig.AddData(_Config)" + row_separator)

            output_lang_file.write("    " + alldata[a] + row_separator)
			
        if output_lang_file != "":
            output_lang_file.write("end" + row_separator)
            output_lang_file.write("return _LandConfig" + row_separator)
            output_lang_file.close()
            output_lang_file = ""
	
    #output_file.write(column_separator)
    output_file.write(row_separator) 
    output_file.write('_Config._length = %s' % (real_count))   
    output_file.write(row_separator)
	
    #output_file.write(row_separator) 
    output_file.write(row_separator) 
    output_file.write('function _Config.getData(Id)')
    output_file.write(row_separator) 
    output_file.write('    local _data = _Config._Data["id_"..Id]')
    output_file.write(row_separator) 
    output_file.write('    if _data then return _data end')
    output_file.write(row_separator) 
    output_file.write('    return nil')
    output_file.write(row_separator)
    output_file.write('end')
	
    output_file.write(row_separator) 
    output_file.write(row_separator) 
    output_file.write('function _Config.getItem(Id, Key)')
    output_file.write(row_separator) 
    output_file.write('    local _data = _Config.getData(Id)')
    output_file.write(row_separator) 
    output_file.write('    if _data then return _data[Key] end')
    output_file.write(row_separator) 
    output_file.write('    return nil')
    output_file.write(row_separator) 
    output_file.write('end')
	
    output_file.write(row_separator) 
    output_file.write(row_separator) 
    output_file.write('function _Config.Data()')
    output_file.write(row_separator) 
    output_file.write('    local _dataList = {}')
    output_file.write(row_separator) 
    output_file.write('    for k,_data in pairs(_Config._Data) do')
    output_file.write(row_separator) 
    output_file.write('        if type(_data) == "table" then table.insert(_dataList, _data) end')
    output_file.write(row_separator) 
    output_file.write('    end')
    output_file.write(row_separator) 
    output_file.write('    return _dataList')
    output_file.write(row_separator) 
    output_file.write('end')

    output_file.write(row_separator)
    output_file.write(row_separator)
    output_file.write('return _Config')
    output_file.close()


def chuli_mubiao(target, output_file):
    if os.path.isfile(target):
        chuli_wenjia(target, output_file)

    if os.path.isdir(target):
        chuli_wenjiajia(target, output_file)


def chuli_wenjiajia(target, output_file):
    for root, dirs, files in os.walk(target):
        for each_file in files:
            if not each_file.startswith('~'):
                file_name = '%s/%s' % (root, each_file)
                chuli_wenjia(file_name, output_file)


def chuli_wenjia(target, output_file):
    (path_name, file_name) = os.path.split(target)
    (base_name, ext) = os.path.splitext(file_name)
    if ext == input_ext:
        duqu_gongzuobo(path_name, base_name, ext, output_file)
    else:
        if ext == input_ext2:
            duqu_gongzuobo(path_name, base_name, ext, output_file)


def main(targets):
    if not len(targets) == 0:
        output_file = codecs.open('%s/%s.lua' % (output_folder, 'WDConfig'), 'w', output_encoding)
        output_file.write("-- Auto Generate, Don't try to modify!")
        output_file.write(row_separator)
        output_file.write(row_separator)
        output_file.write('local _Config = {}')
        output_file.write(row_separator)
        output_file.write(row_separator)
		
        for target in targets:
            chuli_mubiao(target, output_file)
			
        output_file.write(row_separator)
        output_file.write('return _Config')
        output_file.close()
    else:
        print 'need arguments for target'


if __name__ == '__main__':
    main(sys.argv[1:])
